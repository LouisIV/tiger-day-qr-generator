"""
2018
Louis Lombardo IV
LouisLombardoIV@gmail.com
L_lombarod@u.pacific.edu

Creation of QR codes from Excel documents for:
TIGER CREATIVE, ADMISSIONS
"""

from openpyxl import Workbook
# from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# from openpyxl.compat import range
# from openpyxl.utils import get_column_letter


import qrcode
import csv
import sys

# wb = Workbook()
# ws = wb.active

# ws[]

# def loadExcelFile(excel_file):
# 	wb = load_workbook(filename=excel_file)
# 	sheet_ranges = wb['range names']

# 	img = Image('logo.png')
# 	wb.save('logo.xlsx')

def writeRowToFile(ws, data, current_row, export_location):

	# print("")
	# print(data['E'] + ":")
	columns = ['A', 'B', 'C', 'D', 'E']
	for column in columns:
		
		# print("     %s%s : %s" % (column, current_row, data[column]))
		ws["%s%s" % (column, current_row)] = data[column]
		
	location = export_location + "/images/" + data['E'] + ".jpeg"
	qr_img = qrcode.make(data['E'])
	qr_img.save(location)

	img = Image(location)
	ws.add_image(img, ('F%s' % current_row))

	'''
	# Last Name
	ws['A%s' % current_row] = last_name

	# First Name
	ws['B%s' % current_row] = first_name

	# Major
	ws['C%s' % current_row] = major

	# City
	ws['D%s' % current_row] = city

	# ID
	ws['E%s' % current_row] = id_num

	# QR
	ws['F%s' % current_row] = img
	'''

def create_workbook():
	wb = Workbook()
	ws = wb.active
	return wb, ws

def process_csv(csv_location, export_location, export_filename):

	wb, ws = create_workbook()

	with open(csv_location, 'r') as csvfile:

		# Students.CSV
		# +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
		# | Term_Descrip | PACIFIC_ID | LAST_NAME | FIRST_NAME | STREET_LINE1 | ... | PAC_EMAIL |
		# |--------------+------------+-----------+------------+--------------+-----+-----------|
		# | Fall 2018    | 9893000000 | Smith     | John       | N / A        | ... | N / A     |
		#                  |            |           |
		#				   |            |           + > Second part of file name.
		#				   |            + > First part of file name.
		#                  + > Data for the QR generator. 
		# Smith_John.png { 9893000000 }

		current_row = 0
		csv_reader = csv.reader(csvfile, delimiter=',', quotechar='|')
		for row in csv_reader:

			current_row += 1

			if None in (row[2], row[3], row[11], row[7], row[1]):
				raise Exception('One of the values was missing in the CSV file ROW: ', current_row)
			else:
				data = {
					'A': row[2],
					'B': row[3],
					'C': row[15],
					'D': row[7],
					'E': row[1]
				}
				writeRowToFile(ws, data, current_row, export_location)

			if current_row > 250:
				break

	wb.save(export_location + "/" + export_filename)


if __name__ == "__main__":
	if len(sys.argv) > 3:
		print("CSV Location     : %s" % sys.argv[1])
		print("Export Location  : %s" % sys.argv[2])
		print("Export filename  : %s" % sys.argv[3])

		# start_row = False
		# if len(sys.argv) > 3:
		# 	print("Start Row: %s" % sys.argv[3])
		# 	start_row = sys.argv[3]

		# end_row = False
		# if len(sys.argv) > 4:
		# 	print("End Row: %s" % sys.argv[4])
		# 	end_row = sys.argv[4]

		process_csv(sys.argv[1], sys.argv[2], sys.argv[3])
	else:
		print("Please provide a path to your csv and an export location.")

